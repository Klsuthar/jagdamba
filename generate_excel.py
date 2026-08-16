"""
Generate Student_List_All_Classes.xlsx for Jagdamba School 2026-27 session.
Reads promoted student JSON files and creates an Excel workbook with one sheet per class.
"""
import json
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_DIR = os.path.join(BASE_DIR, 'json')
DATA_DIR = os.path.join(BASE_DIR, 'data')
OUTPUT_FILE = os.path.join(DATA_DIR, 'Student_List_All_Classes.xlsx')

# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)

# Column headers
HEADERS = ['roll_no', 'student_name', 'father_name', 'mother_name', 'dob', 'image']

wb = openpyxl.Workbook()
# Remove default sheet
wb.remove(wb.active)

for class_num in range(1, 6):
    sheet_name = f'Class {class_num}'
    ws = wb.create_sheet(title=sheet_name)
    
    # Write headers
    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Read student JSON
    json_file = os.path.join(JSON_DIR, f'class{class_num}', f'class{class_num}_students.json')
    
    if os.path.exists(json_file):
        with open(json_file, 'r', encoding='utf-8') as f:
            students = json.load(f)
        
        for row_idx, student in enumerate(students, 2):
            ws.cell(row=row_idx, column=1, value=student.get('roll_no', ''))
            ws.cell(row=row_idx, column=2, value=student.get('student_name', ''))
            ws.cell(row=row_idx, column=3, value=student.get('father_name', ''))
            ws.cell(row=row_idx, column=4, value=student.get('mother_name', ''))
            ws.cell(row=row_idx, column=5, value=student.get('dob', ''))
            ws.cell(row=row_idx, column=6, value=student.get('image', ''))
        
        print(f"  {sheet_name}: {len(students)} students added")
    else:
        print(f"  {sheet_name}: No JSON file found (empty sheet)")

wb.save(OUTPUT_FILE)
print(f"\nExcel file saved: {OUTPUT_FILE}")
print("Done!")
